import logging
from typing import List, Dict, Optional
from openpyxl import Workbook, load_workbook
from odoo.addons.my_module.model import impression_brother_cmd

logger = logging.getLogger(__name__)


def read_excel(file_path: str) -> List[Dict[str, Optional[str]]]:
    """
    Lit un fichier Excel et retourne les données sous forme de liste de dictionnaires.
    Chaque dictionnaire représente une ligne avec les en-têtes comme clés.
    """
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    if not headers:
        raise ValueError("Aucune en-tête trouvée dans le fichier Excel.")

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data.append(row_data)

    return data


def write_excel(file_path: str, data: List[Dict[str, Optional[str]]]) -> None:
    """
    Écrit des données dans un fichier Excel.
    Les données doivent être une liste de dictionnaires avec les mêmes clés.
    """
    if not data:
        raise ValueError("Les données à écrire sont vides.")

    workbook = Workbook()
    sheet = workbook.active

    headers = list(data[0].keys())
    sheet.append(headers)

    for item in data:
        sheet.append([item.get(header) for header in headers])

    workbook.save(file_path)


def print_sections(sections: Optional[List[tuple]], printer_path: Optional[str] = None) -> None:
    """
    Imprime les sections si disponibles. Gère les erreurs sans interruption du programme.
    """
    if not sections:
        logger.info("Aucune section à imprimer.")
        return

    try:
        impression_brother_cmd.print_postes_with_cut(sections, printer_path=printer_path)
    except Exception as e:
        logger.exception("Erreur lors de l'impression des postes : %s", e)
